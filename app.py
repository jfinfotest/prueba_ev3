import openpyxl

workbook = openpyxl.load_workbook("asistencia.xlsx")
sheet = workbook["Asistencia"]

last_row = sheet.max_row
sheet.insert_rows(last_row + 1)

nombre = input("Ingrese su nombre: ")
fecha = input("Ingrese la fecha (AAAA-MM-DD): ")
hora = input("Ingrese la hora de entrada (HH:MM): ")

sheet.cell(row=last_row + 1, column=1).value = nombre
sheet.cell(row=last_row + 1, column=2).value = fecha
sheet.cell(row=last_row + 1, column=3).value = hora

workbook.save("asistencia.xlsx")
